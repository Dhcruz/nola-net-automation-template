import pandas as pd
from netmiko.ssh_dispatcher import ConnectHandler
from netmiko import SSHDetect,ConnectHandler
from openpyxl import load_workbook
import os


def Leer_dispositivos(archivo):
        # Leer dispositivos desde un archivo
        dispositivos = []
        wb = load_workbook(filename=archivo, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] and row[1] and row[2]:
                dispositivo = {
                    'device_type':'autodetect',
                    'ip': row[0].strip(),
                    'username': row[1].strip(),
                    'password': row[2].strip(),
                    'port': 23
        }
                dispositivos.append(dispositivo)
        return dispositivos
    
def obtener_informacion_equipo(dispositivo):    
        
        try:
            # Conectarse al dispositivo Cisco
            conexion = ConnectHandler(**dispositivo)
            guesser = SSHDetect(**dispositivo)
            best_match = guesser.autodetect()
            # Obtener el nombre de host del dispositivo
            hostname = conexion.find_prompt().replace('#', '')        
            
            # Detección básica del tipo de dispositivo
            if 'cisco_wlc' in best_match:
                comando_inventario = 'show udi'
                comando_version = 'Show boot'
            elif 'cisco_xe' in best_match:
                comando_inventario = 'show inventory | i SN'
                comando_version = 'show version | i , Version'
            elif 'cisco_nxos' in best_match:
                comando_inventario = 'show inventory | i SN'
                comando_version = 'show version | i System'
            elif 'cisco_asa' in best_match:
                comando_inventario = 'show inventory'
                comando_version = 'show version'
            elif 'cisco_ftd' in best_match:
                comando_inventario = 'show inventory'
                comando_version = 'show version'
            elif 'cisco_viptela' in best_match:
                comando_inventario = 'show inventory | i SN'
                comando_version = 'show version | i ersion'
            elif 'cisco_xr' in best_match:
                comando_inventario = 'show inventory | i SN'
                comando_version = 'show version | i Software'
            else:
                return {"host": hostname, "error": "Tipo de dispositivo no soportado"}
            
            # Ejecuta el comando de inventario
            salida_inventory = conexion.send_command(comando_inventario)
            modelo = salida_inventory
            salida_version = conexion.send_command(comando_version)
            version = salida_version

            # Cerrar la conexión
            conexion.disconnect()

            return hostname, modelo, version
        
        except Exception as e:
            print(f"Error al conectar al dispositivo: {e}")
            return None

    #Archivo de lista de dispositivos, Incluya su archivo excel de lista de dispositivos
archivo_dispositivos = 'Lista.xlsx'

#Obtener la lista
lista_dispositivos = Leer_dispositivos(archivo_dispositivos)

for dispositivo in lista_dispositivos:
    hostname, modelo, version = obtener_informacion_equipo(dispositivo)

    if hostname and modelo and version:
        # Crear un DataFrame con los datos
        data = {'hostname': [hostname], 'ip': [dispositivo['ip']], 'modelo': [modelo], 'version': [version]}
        df = pd.DataFrame(data)

        #Archivo de inventario, Incluya su archivo excel donde se guardará la información extraída
        archivo_excel = "inventario.xlsx"
        ruta_completa = os.path.join(os.getcwd(), archivo_excel)
        if os.path.exists(ruta_completa):
            df_existente = pd.read_excel(ruta_completa)
            df = pd.concat([df_existente, df])

        # Guardar el DataFrame actualizado en el archivo Excel
        df.to_excel(ruta_completa, index=False)
        print("Información guardada exitosamente en el archivo 'inventario.xlsx'")
    else:
        print("No se pudo obtener la información del dispositivo.")
